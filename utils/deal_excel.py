#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @File  : deal_excel.py
# @Author: itnoobzzy
# @Date  : 2021/3/16
# @Desc  : excel 操作

from openpyxl import Workbook


class DealExcel:
    """excel的读写操作"""

    def __init__(self):
        self.xlsxBook = Workbook()

    def create_sheet(self, title, index, **kwargs):
        """获取sheet对象"""
        sheet = self.xlsxBook.create_sheet(title, index)
        for key, value in kwargs.items():
            sheet[key] = value
        return sheet

    def save(self, file_name):
        self.xlsxBook.save(file_name)

